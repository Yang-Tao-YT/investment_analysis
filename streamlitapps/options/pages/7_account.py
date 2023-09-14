import datetime
import os
import time

import pandas as pd
import streamlit as st

import option_strategy
from data.generate_data import DataLoader
from strategy.factor_cross_section import main as cross_section
import pathlib
import sys
from utils.fund_records import main
sys.path.append(str(pathlib.Path().absolute()).split("/src")[0] )


import datetime
import re
from itertools import groupby

from openpyxl import Workbook, load_workbook

st.set_page_config(
page_title="investing analysis",  #页面标题
page_icon=":rainbow:",  #icon
layout="wide", #页面布局
initial_sidebar_state="auto" #侧边栏
)

debug = 0

worbok = load_workbook(f'../finance_data/options.xlsx', data_only=True)
# worbok = pd.read_excel(f'../finance_data/options.xlsx', sheet_name='净值', )
st.dataframe(pd.DataFrame(worbok['净值'].values).iloc[-10: , :10])

# worbok['净值'][f"A{str_}"].value
if st.checkbox('修改') or debug:
    worbok = load_workbook(f'../finance_data/options.xlsx',)
    print(worbok['净值']["1:1"])
    print([i.value for i in worbok['净值']["1:1"]])
    str_= len(worbok['净值']["A:A"]) + 1

    amount = st.number_input('净值') if not debug else 1357218.4
    mem = st.text_input('备注', value = '')
    risk = st.text_input('风险事件', value = '')
    if st.button('确认修改') or debug:
        worbok['净值'][f"B{str_}"].value = amount
        if mem != '':
            worbok['净值'][f"I{str_}"].value = mem
        if risk != '':
            worbok['净值'][f"J{str_}"].value = risk
        main(worbok)

        

