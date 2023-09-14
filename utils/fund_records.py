
from openpyxl import Workbook, load_workbook
import re
from itertools import groupby
import datetime

def plus1(func):
	func = [''.join(list(g)) for k,g in groupby(func, lambda x: x.isdigit())]
	result = []
	for count, i in enumerate(func):
		if count == 0:# 第一个直接加
			result += [i]
			continue
			
		if i.isdigit() and func[count - 1][-1].isalpha() :# 如果为数字检查上一个尾是否为字母，若是则i是单元格位置
			if func[count - 1][-2] != '$': #判断是不是固定单元格
				result += [str(int(i) + 1)]
			else:
				result += [i]
			continue


		result += [i]

	func = ''.join(result)
	return func

def pulldown(worbok, sheetname, row, columns):
	if isinstance(worbok[sheetname][f'{columns}{row-1}'].value, str):
		if '=' in worbok[sheetname][f'{columns}{row-1}'].value:
			worbok[sheetname][f'{columns}{row}'].value = plus1(worbok[sheetname][f'{columns}{row-1}'].value)

			worbok[sheetname][f'{columns}{row}'].number_format = worbok[sheetname][f'{columns}{row-1}'].number_format
	
		else:
			worbok[sheetname][f'{columns}{row}'].value = worbok[sheetname][f'{columns}{row-1}'].value

	elif  isinstance(worbok[sheetname][f'{columns}{row - 1}'].value, datetime.datetime):
			worbok[sheetname][f'{columns}{row}'].value = worbok[sheetname][f'{columns}{row-1}'].value + datetime.timedelta(1)
			worbok[sheetname][f'{columns}{row}'].number_format = worbok[sheetname][f'{columns}{row-1}'].number_format

	else:
		worbok[sheetname][f'{columns}{row}'].value = worbok[sheetname][f'{columns}{row-1}'].value


# worbok['净值']['C52'].value = 1
def main(worbok):
	for _sheetname in ['杨文萍固收', '杨翠红固收', '陶春强', '陶阳', '杨文萍', 'sum']:
		print(worbok[_sheetname]["1:1"])
		print([i.value for i in worbok[_sheetname]["1:1"]])
		str_= len(worbok[_sheetname]["A:A"]) + 1
		for ix in ['A', 'B', 'C', 'D', 'E', 'F' ,'G']:
			pulldown(worbok, _sheetname, str_, ix)

	for _sheetname in ['净值']:
		print(worbok[_sheetname]["1:1"])
		print([i.value for i in worbok[_sheetname]["1:1"]])
		str_= len(worbok[_sheetname]["A:A"]) 
		for ix in ['A', 'C', 'D', 'E', 'F' ,'G']:
			if ix == 'A':
				print(1)
			pulldown(worbok, _sheetname, str_, ix)


	# worbok['净值'][f'E{str_}'].number_format = '0.00%'
	worbok.save(f'../finance_data/options.xlsx') 

if __name__ == '__main__':
	pass
	# worbok = load_workbook(f'../finance_data/options.xlsx')
	# worbok.sheetnames

